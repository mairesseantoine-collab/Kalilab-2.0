#!/usr/bin/env python3
"""
Génère un fichier Excel de test qui reproduit la structure attendue
de l'onglet ANALYSE (deux blocs STE et STM).

Usage :  python iq20_generate_test_data.py
Sortie : IQ20_TEST.xlsx  (dans le répertoire courant)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ANALYSE"

MOIS = ["Juil 25", "Août 25", "Sept 25", "Oct 25", "Nov 25", "Déc 25"]

# Données STE (réalistes)
ste_data = {
    "Demandes A":                       [1420, 1380, 1510, 1600, 1450, 1390],
    "Demandes A non vérifiées":         [ 284,  207,  302,  288,  261,  250],
    "Demandes A vérifiées":             [1136, 1173, 1208, 1312, 1189, 1140],
    "Demandes A vérifiées avec erreur": [  91,  106,  157,  118,   95,   91],
    "Demandes A vérifiées sans erreur": [1045, 1067, 1051, 1194, 1094, 1049],
    "Demandes H":                       [ 340,  320,  360,  410,  390,  370],
}

# Données STM (légèrement différentes)
stm_data = {
    "Demandes A":                       [ 980,  960, 1020, 1100, 1010,  970],
    "Demandes A non vérifiées":         [ 196,  144,  204,  198,  182,  175],
    "Demandes A vérifiées":             [ 784,  816,  816,  902,  828,  795],
    "Demandes A vérifiées avec erreur": [  55,   65,   98,   72,   58,   56],
    "Demandes A vérifiées sans erreur": [ 729,  751,  718,  830,  770,  739],
    "Demandes H":                       [ 210,  195,  225,  260,  240,  220],
}

HDR_FILL  = PatternFill("solid", fgColor="1565C0")
SITE_FILL = PatternFill("solid", fgColor="0D47A1")
HDR_FONT  = Font(color="FFFFFF", bold=True)
SITE_FONT = Font(color="FFFFFF", bold=True, size=13)

def write_block(ws, start_row: int, site: str, data: dict):
    # Entête site
    cell = ws.cell(row=start_row, column=1, value=site)
    cell.font = SITE_FONT
    cell.fill = SITE_FILL
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=len(MOIS)+1)

    # Ligne mois
    for j, mois in enumerate(MOIS, start=2):
        c = ws.cell(row=start_row+1, column=j, value=mois)
        c.font  = HDR_FONT
        c.fill  = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    # Lignes de données
    for i, (label, values) in enumerate(data.items(), start=2):
        ws.cell(row=start_row+i, column=1, value=label)
        for j, val in enumerate(values, start=2):
            ws.cell(row=start_row+i, column=j, value=val)

    return start_row + len(data) + 3   # prochaine ligne disponible

row = 1
row = write_block(ws, row, "STE", ste_data)
row = write_block(ws, row, "STM", stm_data)

ws.column_dimensions["A"].width = 36
for col in "BCDEFG":
    ws.column_dimensions[col].width = 12

path = "IQ20_TEST.xlsx"
wb.save(path)
print(f"[OK] Fichier de test généré : {path}")
print("Lancez maintenant : python iq20_analyse.py IQ20_TEST.xlsx")
